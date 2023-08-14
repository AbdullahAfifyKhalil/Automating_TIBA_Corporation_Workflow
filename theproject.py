# Import Needed Libraries
import re
import sys
import pandas as pd
from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtGui import QFont, QTextCursor
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QTextEdit
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class MaterialCostCalculation(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Material Cost Calculation")
        self.setGeometry(100, 100, 600, 600)

        icon = QtGui.QIcon("logo2.png") 
        self.setWindowIcon(icon)

        self.header_image = QtWidgets.QLabel(self)
        self.header_image.setGeometry(20, 20, 400, 50)
        
        header_pixmap = QtGui.QPixmap("company_icon.png")
        self.header_image.setPixmap(header_pixmap)

        header_pixmap = header_pixmap.scaled(self.header_image.size(), QtCore.Qt.AspectRatioMode.KeepAspectRatio)
        self.header_image.setPixmap(header_pixmap)
        
        self.boq_file_label = QtWidgets.QLabel(self)
        self.boq_file_label.setText("BOQ Sheet:")
        self.boq_file_label.move(20, 90)
        
        self.boq_file_entry = QtWidgets.QLineEdit(self)
        self.boq_file_entry.setGeometry(20, 120, 400, 30)
        
        self.browse_boq_file_button = QtWidgets.QPushButton(self)
        self.browse_boq_file_button.setText("Browse")
        self.browse_boq_file_button.setGeometry(430, 120, 100, 30)
        self.browse_boq_file_button.clicked.connect(self.browse_boq_file)
        
        self.reference_sheet_file_label = QtWidgets.QLabel(self)
        self.reference_sheet_file_label.setText("Reference Sheet:")
        self.reference_sheet_file_label.move(20, 160)
        
        self.reference_sheet_file_entry = QtWidgets.QLineEdit(self)
        self.reference_sheet_file_entry.setGeometry(20, 190, 400, 30)
        
        self.browse_reference_sheet_file_button = QtWidgets.QPushButton(self)
        self.browse_reference_sheet_file_button.setText("Browse")
        self.browse_reference_sheet_file_button.setGeometry(430, 190, 100, 30)
        self.browse_reference_sheet_file_button.clicked.connect(self.browse_reference_sheet_file)
        
        self.process_button = QtWidgets.QPushButton(self)
        self.process_button.setText("Process Files")
        self.process_button.setGeometry(20, 250, 100, 30)
        self.process_button.clicked.connect(self.process_files)
        
        self.output_text_label = QtWidgets.QLabel(self)
        self.output_text_label.setText("Processed Data:")
        self.output_text_label.setGeometry(20, 290, 200, 30)
        
        self.output_text = QtWidgets.QTextEdit(self)
        self.output_text.setGeometry(20, 320, 560, 230)
        
        self.workbook = Workbook()
        
    def browse_boq_file(self):
        boq_file_path, _ = QFileDialog.getOpenFileName(self, "Select BOQ File", "", "Excel Files (*.xlsx)")
        self.boq_file_entry.setText(boq_file_path)
        
    def browse_reference_sheet_file(self):
        reference_sheet_file_path, _ = QFileDialog.getOpenFileName(self, "Select Reference Sheet File", "", "Excel Files (*.xlsx)")
        self.reference_sheet_file_entry.setText(reference_sheet_file_path)
        
    def process_files(self):
        boq_file_path = self.boq_file_entry.text()
        reference_sheet_file_path = self.reference_sheet_file_entry.text()
        
        try:
            boq_data = pd.read_excel(boq_file_path)
            reference_sheet_data = pd.read_excel(reference_sheet_file_path)
            
            # Identify the columns containing the material, volume, and unit information in the BOQ file
            material_column1 = None
            volume_column = None
            unit_column = None

            for column in boq_data.columns:
                for value in boq_data[column].astype(str).str.upper().values:
                    if 'MATERIAL' in value:
                        material_column1 = column
                    elif 'VOLUME' in value:
                        volume_column = column
                    elif 'UNIT' in value:
                        unit_column = column

                if material_column1 is not None and volume_column is not None and unit_column is not None:
                    break

            if material_column1 is None or volume_column is None or unit_column is None:
                raise ValueError("Material, volume, or unit column not found in BOQ file.")

            # Remove the bold text items from the material column
            bold_pattern = r'\*\*([^*]+)\*\*'
            df_all_materials = boq_data.copy()
            df_all_materials[material_column1] = df_all_materials[material_column1].apply(lambda x: re.sub(bold_pattern, r'\1', str(x)))

            # Filter out rows with empty material values
            df_all_materials = df_all_materials[df_all_materials[material_column1].str.strip() != ""]

            # Remove rows with NaN in the volume column
            df_all_materials = df_all_materials.dropna(subset=[volume_column])

            # Create the final df_all_materials DataFrame with the filtered material items
            df_all_materials = df_all_materials[[material_column1, volume_column, unit_column]].copy()
            df_all_materials.columns = ["Material", "Volume", "Unit"]

            # Create separate lists for each material category
            material_lists = {}
            current_category = None

            for index, row in boq_data.iterrows():
                value = row[material_column1]
                volume = row[volume_column]
                unit = row[unit_column]


            # Extract the material, price IDR, and price USD columns
            material_price_data = reference_sheet_data[['MATERIAL', 'PRICE_IDR', 'PRICE_USD', 'DELIVERY TIME (DAYS)']].copy()

            # Compare materials and create a new DataFrame with matched information
            matched_data = df_all_materials.merge(material_price_data, left_on='Material', right_on='MATERIAL', how='inner')

            # Calculate total IDR and total USD
            matched_data['Total IDR'] = matched_data['Volume'] * matched_data['PRICE_IDR']
            matched_data['Total USD'] = matched_data['Volume'] * matched_data['PRICE_USD']

            # Identify non-matched material items
            non_matched_materials = df_all_materials[~df_all_materials['Material'].isin(matched_data['Material'])]

            # Create a new DataFrame with one material column and other columns as discussed earlier
            final_data = matched_data[['Material', 'Volume', 'Unit', 'PRICE_IDR', 'PRICE_USD', 'Total IDR', 'Total USD']]

            # Calculate the sum of Total IDR and Total USD
            total_idr_sum = final_data['Total IDR'].sum()
            total_usd_sum = final_data['Total USD'].sum()

            # Create sum rows as a DataFrame
            sum_row = pd.DataFrame({'Material': ['Total'], 'Volume': [None], 'Unit': [None], 'PRICE_IDR': [None], 'PRICE_USD': [None], 'Total IDR': [total_idr_sum], 'Total USD': [total_usd_sum]})

            # Concatenate sum rows with the final_data DataFrame
            final_data_with_sum = pd.concat([final_data, sum_row], ignore_index=True)

            # Sort the final_data DataFrame by the Material column in ascending order
            final_data_with_sum = final_data_with_sum.sort_values(by='Material', ascending=True)

            # Select the Material and DELIVERY TIME (DAYS) columns
            schedule_data = matched_data[['MATERIAL', 'DELIVERY TIME (DAYS)']]

            # Calculate the start week, end week, and number of weeks
            schedule_data['START WEEK'] = (schedule_data['DELIVERY TIME (DAYS)'] // 7) + 1
            schedule_data['END WEEK'] = schedule_data['START WEEK'] + schedule_data['DELIVERY TIME (DAYS)'] % 7
            schedule_data['NUM OF WEEKS'] = schedule_data['DELIVERY TIME (DAYS)'] // 7 + 1

            # Sort the data by the 'Start Week' column in ascending order
            schedule_data = schedule_data.sort_values(by='START WEEK', ascending=True)

            # Get the maximum number of weeks
            max_weeks = int(schedule_data['END WEEK'].max())

            # Create a new Excel workbook
            workbook = Workbook()
            sheet = workbook.active

            # Write the column headers
            headers = ['Material', 'Delivery Time (Days)', 'Start Week', 'End Week', 'Num of Weeks']
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col, value=header).font = Font(bold=True)

            # Create a grayscale gradient color palette
            materials = schedule_data['MATERIAL'].unique()
            num_materials = len(materials)
            start_color = (204, 204, 204)  # RGB value for the starting color (light gray)
            end_color = (0, 0, 0)  # RGB value for the ending color (black)
            color_palette = []
            for i in range(num_materials):
                r = start_color[0] - i * (start_color[0] - end_color[0]) // (num_materials - 1)
                g = start_color[1] - i * (start_color[1] - end_color[1]) // (num_materials - 1)
                b = start_color[2] - i * (start_color[2] - end_color[2]) // (num_materials - 1)
                color_palette.append(f"FF{r:02X}{g:02X}{b:02X}")  # Convert RGB to hex format

            week_start_column = 6
            for week in range(max_weeks):
                week_column = week_start_column + week
                sheet.cell(row=1, column=week_column).value = f"Week {week + 1}"
                sheet.column_dimensions[get_column_letter(week_column)].width = 10

            for row, row_data in enumerate(schedule_data.itertuples(index=False), start=2):
                material, delivery_time, start_week, end_week, num_of_weeks = row_data

                num_of_weeks = int(num_of_weeks)

                sheet.cell(row=row, column=1, value=material)
                sheet.cell(row=row, column=2, value=delivery_time)
                sheet.cell(row=row, column=3, value=start_week)
                sheet.cell(row=row, column=4, value=end_week)
                sheet.cell(row=row, column=5, value=num_of_weeks)

                fill = PatternFill(fill_type='solid', fgColor=color_palette[row - 2])
                for week in range(num_of_weeks):
                    week_column = week_start_column + start_week - 1 + week
                    if week_column <= max_weeks + week_start_column - 1:
                        sheet.cell(row=row, column=week_column).fill = fill
            
            # Display the processed data  .........................
            self.output_text.clear()
            self.output_text.insertPlainText("BOQ Data:\n" + str(df_all_materials) + "\n\n")
            self.output_text.insertPlainText("Matched Data:\n" + str(final_data_with_sum) + "\n\n")
            self.output_text.insertPlainText("Non-Matched Materials:\n" + str(non_matched_materials) + "\n\n")
            self.output_text.insertPlainText("Construction Schedule:\n" + str(schedule_data))
            
            # Save the output Excel files
            output_file_path = "Material Cost Calculation.xlsx"
            final_data_with_sum.to_excel(output_file_path, index=False)
            
            non_matched_file_path = "Requesting Quotes.xlsx"
            non_matched_materials.to_excel(non_matched_file_path, columns=["Material", "Volume", "Unit"], index=False)
            
            output_file_pat = "Construction Schedule.xlsx"
            workbook.save(output_file_pat)
            
            # Show success message
            QMessageBox.information(self, "Success", "Files processed successfully. Output files saved.")
            
        except Exception as e:
            # Show error message
            QMessageBox.critical(self, "Error", str(e))
    
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = MaterialCostCalculation()
    window.show()
    sys.exit(app.exec_())
